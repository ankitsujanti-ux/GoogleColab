# utils.py
import logging
import os
import sys
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime
from typing import Any, Dict, Iterable, List, Optional

import pandas as pd

from config import EXCEL_PATH, THRESHOLD_LOW, THRESHOLD_MED

logger = logging.getLogger(__name__)

_last_mtime = None
_cached_df = None

EXCEL_INVALID_MSG = (
    "The file at EXCEL_PATH is not a valid Excel (.xlsx) file. "
    "It may be corrupted, empty, or saved in another format (e.g. CSV). "
    "Please ensure the file in .env (EXCEL_PATH) is a valid .xlsx workbook."
)

# Required columns for high-risk detection, notifications, and clinician routing.
# If missing, they are added to the Excel file with default empty/NaN values.
REQUIRED_EXCEL_COLUMNS = [
    "Member ID",
    "Patient Name",
    "Adherence Percentage",
    "Adeherence Percentage",
    "PDC Percentage",
    "Refill Days",
    "Patient Contact",
    "Patient EmailID",
    "City",
    "Medication Name",
    "EventDate",
    "NotificationSentOn",
    "NotificationSentAt",
    "NotificationStatus",
    "NotificationChannel",
    "NotificationMessageId",
    "RoutedToClinicianOn",
    "RoutedToClinicianNote",
    "AppreciationSentOn",
    "RiskScore",
    "RiskLabel",
    "ClinicianAssigned",
    "EscalationReason",
    "LastUpdated",
]

def ensure_excel_columns(path: Optional[str] = None) -> None:
    """
    Ensure the Excel file has all required columns. Adds any missing columns
    with default empty/NaN values and saves the file so updates work correctly.
    """
    p = path or EXCEL_PATH
    if not os.path.exists(p):
        return
    try:
        raw_df = pd.read_excel(p, engine="openpyxl")
        raw_df.columns = [str(c).strip() for c in raw_df.columns]
        member_col = "Member ID" if "Member ID" in raw_df.columns else ("Member_ID" if "Member_ID" in raw_df.columns else None)
        if member_col is None:
            raw_df["Member ID"] = raw_df.index.astype(str)
        changed = False
        for col in REQUIRED_EXCEL_COLUMNS:
            if col not in raw_df.columns:
                if col in ("NotificationSentOn", "NotificationSentAt", "RoutedToClinicianOn", "AppreciationSentOn", "LastUpdated", "EventDate"):
                    raw_df[col] = pd.NaT
                elif col in ("Adherence Percentage", "Adeherence Percentage", "PDC Percentage", "Refill Days", "RiskScore"):
                    raw_df[col] = pd.NA
                else:
                    raw_df[col] = ""
                changed = True
        if changed:
            raw_df.to_excel(p, index=False, engine="openpyxl")
    except Exception:
        pass


def update_excel_column_values(
    member_ids: Iterable[str],
    column_updates: Dict[str, Any],
    path: Optional[str] = None
) -> None:
    """
    Update specific column values in the Excel file for given Member IDs.
    Automatically adds missing columns if they don't exist.
    
    Args:
        member_ids: List of Member IDs to update
        column_updates: Dictionary mapping column names to values
        path: Optional path to Excel file (defaults to EXCEL_PATH)
    """
    global _cached_df, _last_mtime
    p = path or EXCEL_PATH
    if not os.path.exists(p):
        raise FileNotFoundError(f"Excel file not found at path: {p}")
    
    member_ids = list(member_ids)
    if not member_ids or not column_updates:
        return
    
    try:
        # Ensure columns exist first
        ensure_excel_columns(p)
        
        raw_df = pd.read_excel(p, engine="openpyxl")
        raw_df.columns = [str(c).strip() for c in raw_df.columns]
        
        member_col = "Member ID" if "Member ID" in raw_df.columns else ("Member_ID" if "Member_ID" in raw_df.columns else None)
        if member_col is None:
            raise KeyError("Member ID column not found in the Excel sheet.")
        
        # Add missing columns if needed
        for col in column_updates.keys():
            if col not in raw_df.columns:
                if col in ("NotificationSentOn", "NotificationSentAt", "RoutedToClinicianOn", "LastUpdated", "EventDate"):
                    raw_df[col] = pd.NaT
                elif col in ("Adherence Percentage", "Adeherence Percentage", "PDC Percentage", "Refill Days", "RiskScore"):
                    raw_df[col] = pd.NA
                else:
                    raw_df[col] = ""
        
        # Update values for matching Member IDs
        mask = raw_df[member_col].astype(str).isin([str(m) for m in member_ids])
        for col, value in column_updates.items():
            if col in raw_df.columns:
                raw_df.loc[mask, col] = value
        
        # Update LastUpdated timestamp
        if "LastUpdated" in raw_df.columns:
            raw_df.loc[mask, "LastUpdated"] = pd.Timestamp.now()
        
        raw_df.to_excel(p, index=False, engine="openpyxl")
        
        # Refresh cache
        _cached_df = compute_fields(raw_df.copy())
        _last_mtime = os.path.getmtime(p)
    except Exception as e:
        raise RuntimeError(f"Error updating Excel columns: {str(e)}")


def route_to_clinician(
    member_ids: Iterable[str],
    reason: str,
    clinician_note: Optional[str] = None,
    path: Optional[str] = None
) -> None:
    """
    Route patient records to clinicians by updating the Excel file.
    
    Args:
        member_ids: List of Member IDs to route
        reason: Escalation reason (e.g., "High risk score", "Refill overdue")
        clinician_note: Optional note for the clinician
        path: Optional path to Excel file
    """
    now = pd.Timestamp.now()
    note = clinician_note or f"Escalated: {reason}"
    
    update_excel_column_values(
        member_ids=member_ids,
        column_updates={
            "RoutedToClinicianOn": now,
            "RoutedToClinicianNote": note,
            "EscalationReason": reason,
        },
        path=path
    )

def _sibling_csv_path(path: str) -> str:
    return os.path.splitext(path)[0] + ".csv"


def _purge_broken_openpyxl() -> None:
    """Remove a partially-initialized openpyxl from sys.modules (common on some hosts)."""
    mod = sys.modules.get("openpyxl")
    if mod is None:
        return
    if getattr(mod, "load_workbook", None) is not None and getattr(mod, "__version__", None):
        return
    for key in list(sys.modules):
        if key == "openpyxl" or key.startswith("openpyxl."):
            del sys.modules[key]


def _col_letters_to_index(letters: str) -> int:
    n = 0
    for ch in letters:
        if not ("A" <= ch <= "Z"):
            break
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return max(n - 1, 0)


def _read_xlsx_via_zip(path: str) -> pd.DataFrame:
    """Stdlib-only .xlsx reader (sharedStrings + first sheet). Avoids broken openpyxl installs."""
    ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    with zipfile.ZipFile(path, "r") as zf:
        shared: List[str] = []
        if "xl/sharedStrings.xml" in zf.namelist():
            root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
            for si in root.findall("m:si", ns):
                texts = [t.text or "" for t in si.findall(".//m:t", ns)]
                shared.append("".join(texts))

        sheet_name = "xl/worksheets/sheet1.xml"
        if sheet_name not in zf.namelist():
            sheets = sorted(
                n for n in zf.namelist()
                if n.startswith("xl/worksheets/sheet") and n.endswith(".xml")
            )
            if not sheets:
                return pd.DataFrame()
            sheet_name = sheets[0]

        root = ET.fromstring(zf.read(sheet_name))
        matrix: List[List[Any]] = []
        for row in root.findall("m:sheetData/m:row", ns):
            cells: Dict[int, Any] = {}
            max_idx = -1
            for c in row.findall("m:c", ns):
                ref = c.get("r") or ""
                letters = "".join(ch for ch in ref if ch.isalpha())
                idx = _col_letters_to_index(letters) if letters else (max_idx + 1)
                max_idx = max(max_idx, idx)
                cell_type = c.get("t")
                v_node = c.find("m:v", ns)
                if v_node is None or v_node.text is None:
                    val: Any = None
                elif cell_type == "s":
                    try:
                        val = shared[int(v_node.text)]
                    except Exception:
                        val = v_node.text
                elif cell_type == "inlineStr":
                    is_node = c.find("m:is", ns)
                    val = "".join(t.text or "" for t in (is_node.findall(".//m:t", ns) if is_node is not None else []))
                else:
                    raw = v_node.text
                    try:
                        val = float(raw) if raw is not None and "." in raw else int(raw) if raw is not None else None
                    except Exception:
                        val = raw
                cells[idx] = val
            if max_idx < 0:
                matrix.append([])
            else:
                matrix.append([cells.get(i) for i in range(max_idx + 1)])

    if not matrix:
        return pd.DataFrame()
    width = max(len(r) for r in matrix)
    matrix = [r + [None] * (width - len(r)) for r in matrix]
    headers = [
        str(h).strip() if h is not None and str(h).strip() else f"col_{i}"
        for i, h in enumerate(matrix[0])
    ]
    data = [dict(zip(headers, row)) for row in matrix[1:]]
    return pd.DataFrame(data)


def _read_excel_dataframe(path: str) -> pd.DataFrame:
    """
    Read patient data robustly.
    Order: direct CSV path → pandas/openpyxl → load_workbook → stdlib zip → sibling CSV.
    Sibling CSV exists for hosts where openpyxl imports are broken (Render).
    """
    csv_path = _sibling_csv_path(path)
    errors: List[str] = []

    if path.lower().endswith(".csv"):
        return pd.read_csv(path, low_memory=False)

    if os.path.isfile(path):
        _purge_broken_openpyxl()
        try:
            import openpyxl  # noqa: F401
            if not getattr(openpyxl, "__version__", None):
                openpyxl.__version__ = "3.1.5"
            return pd.read_excel(path, engine="openpyxl")
        except Exception as first_err:
            errors.append(f"pandas/openpyxl: {first_err}")
            logger.warning("pandas read_excel failed (%s); trying alternate readers", first_err)
            _purge_broken_openpyxl()

        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
            if not rows:
                return pd.DataFrame()
            headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
            data = [dict(zip(headers, row)) for row in rows[1:]]
            return pd.DataFrame(data)
        except Exception as second_err:
            errors.append(f"load_workbook: {second_err}")
            logger.warning("openpyxl load_workbook failed (%s); trying CSV/zip fallbacks", second_err)
            _purge_broken_openpyxl()

    # Prefer known-good CSV before the stdlib zip reader (zip headers can be wrong
    # for some workbooks; CSV is shipped alongside patients.xlsx for Render).
    if os.path.isfile(csv_path):
        logger.warning("Falling back to CSV patient data at %s", csv_path)
        return pd.read_csv(csv_path, low_memory=False)

    if os.path.isfile(path):
        try:
            df = _read_xlsx_via_zip(path)
            if not df.empty:
                logger.info("Loaded Excel via stdlib zip reader (%d rows)", len(df))
                return df
            errors.append("zip reader returned empty")
        except Exception as zip_err:
            errors.append(f"zip: {zip_err}")
            logger.warning("stdlib zip xlsx reader failed (%s)", zip_err)

    raise RuntimeError(
        "Failed to read patient data at {0} (csv={1}): {2}".format(
            path, csv_path, " | ".join(errors) or "no readable source"
        )
    )


def load_patient_data(force_reload: bool = False) -> pd.DataFrame:
    """
    Loads Excel data and caches it; reloads automatically when file time changes.
    Uses EXCEL_PATH from config (set via .env). Ensures required columns exist
    before loading so schema changes are handled and high-risk detection keeps working.
    Returns empty DataFrame if file is missing (production-friendly: app still starts).
    Falls back to sibling patients.csv when openpyxl cannot read the .xlsx (e.g. Render).
    """
    global _last_mtime, _cached_df

    csv_path = _sibling_csv_path(EXCEL_PATH)
    if not os.path.exists(EXCEL_PATH) and not os.path.exists(csv_path):
        _cached_df = pd.DataFrame()
        _last_mtime = None
        return pd.DataFrame()
    try:
        if os.path.exists(EXCEL_PATH) and EXCEL_PATH.lower().endswith((".xlsx", ".xlsm")):
            ensure_excel_columns(EXCEL_PATH)
    except Exception as e:
        logger.warning("ensure_excel_columns skipped: %s", e)
    mtimes = []
    if os.path.exists(EXCEL_PATH):
        mtimes.append(os.path.getmtime(EXCEL_PATH))
    if os.path.exists(csv_path):
        mtimes.append(os.path.getmtime(csv_path))
    mtime = max(mtimes) if mtimes else None
    if force_reload or _cached_df is None or _last_mtime != mtime:
        try:
            source = EXCEL_PATH if os.path.exists(EXCEL_PATH) else csv_path
            df = _read_excel_dataframe(source)
        except Exception as e:
            logger.error("Failed to read Excel at %s: %s", EXCEL_PATH, e)
            raise
        df.columns = [str(c).strip() for c in df.columns]
        for col in ["PDC Percentage", "Adeherence Percentage", "Adherence Percentage", "Refill Days"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")
        df = compute_fields(df)
        _cached_df = df
        _last_mtime = mtime
        logger.info("Loaded %d patients from %s", len(df), os.path.basename(source))
    return _cached_df.copy()

def compute_fields(df: pd.DataFrame) -> pd.DataFrame:
    # prefer Adeherence Percentage if present else PDC Percentage
    adherence_col = None
    if "Adeherence Percentage" in df.columns:
        adherence_col = "Adeherence Percentage"
    elif "Adherence Percentage" in df.columns:
        adherence_col = "Adherence Percentage"
    elif "PDC Percentage" in df.columns:
        df["Adeherence Percentage"] = df["PDC Percentage"]
        adherence_col = "Adeherence Percentage"

    if adherence_col is None:
        adherence_col = "Adeherence Percentage"
        df[adherence_col] = 0

    df[adherence_col] = df[adherence_col].fillna(0)
    # ensure both spellings exist for downstream compatibility
    if adherence_col == "Adherence Percentage" and "Adeherence Percentage" not in df.columns:
        df["Adeherence Percentage"] = df["Adherence Percentage"]
    if adherence_col == "Adeherence Percentage":
        df["Adherence Percentage"] = df["Adeherence Percentage"]

    # classify adherence percentage into Low/Medium/High
    def classify(a):
        if a < THRESHOLD_LOW:
            return "Low"
        if a < THRESHOLD_MED:
            return "Medium"
        return "High"
    df["adherence_group"] = df["Adeherence Percentage"].apply(classify)

    # standard display label for patient: "ID - Name"
    if "Member ID" in df.columns and "Patient Name" in df.columns:
        df["patient_display"] = df["Member ID"].astype(str) + " - " + df["Patient Name"].astype(str)

    # Refill Days = days until refill / days of supply left; adherence based on this column only
    refill_days_col = next((c for c in ["Refill Days", "RefillDays", "Refill_Days"] if c in df.columns), None)
    if refill_days_col:
        rd = pd.to_numeric(df[refill_days_col], errors="coerce")
        df["days_until_refill"] = rd
        df["refill_due"] = (rd < 7) & (rd.notna())
    else:
        df["days_until_refill"] = pd.NA
        df["refill_due"] = False

    # ensure notification and routing columns exist
    if "NotificationSentOn" not in df.columns:
        df["NotificationSentOn"] = pd.NaT
    else:
        df["NotificationSentOn"] = pd.to_datetime(df["NotificationSentOn"], errors="coerce")
    if "NotificationStatus" not in df.columns:
        df["NotificationStatus"] = ""
    if "RoutedToClinicianOn" not in df.columns:
        df["RoutedToClinicianOn"] = pd.NaT
    else:
        df["RoutedToClinicianOn"] = pd.to_datetime(df["RoutedToClinicianOn"], errors="coerce")
    if "RoutedToClinicianNote" not in df.columns:
        df["RoutedToClinicianNote"] = ""

    if "EventDate" in df.columns:
        df["EventDate"] = pd.to_datetime(df["EventDate"], errors="coerce")

    # contact safe string
    for col in ["Patient Contact", "Patient EmailID", "Patient Name"]:
        if col not in df.columns:
            df[col] = None

    # set an ID column if not present
    if "Member ID" not in df.columns:
        df["Member ID"] = df.index.astype(str)

    return df


def update_notification_sent_on(
    member_ids: Iterable[str],
    timestamp: Optional[datetime] = None,
    details_by_member_id: Optional[Dict[str, Dict[str, Any]]] = None,
) -> None:
    """
    Update NotificationSentOn column for provided Member IDs with given timestamp (date).
    Persists changes to the Excel file and refreshes the cached dataframe.
    """
    global _cached_df, _last_mtime

    member_ids = list(member_ids)
    if not member_ids:
        return

    if not os.path.exists(EXCEL_PATH):
        return  # No-op when Excel not configured (e.g. first run)

    # Keep both a date-level and time-level stamp for auditability / dashboard popups
    ts_full = pd.Timestamp(timestamp if timestamp is not None else pd.Timestamp.now())
    ts_date = ts_full.normalize()

    raw_df = pd.read_excel(EXCEL_PATH, engine="openpyxl")
    raw_df.columns = [c.strip() for c in raw_df.columns]

    member_col = None
    if "Member ID" in raw_df.columns:
        member_col = "Member ID"
    elif "Member_ID" in raw_df.columns:
        member_col = "Member_ID"
    else:
        raise KeyError("Member ID column not found in the Excel sheet.")

    if "NotificationSentOn" not in raw_df.columns:
        raw_df["NotificationSentOn"] = pd.NaT
    if "NotificationSentAt" not in raw_df.columns:
        raw_df["NotificationSentAt"] = pd.NaT
    if "NotificationStatus" not in raw_df.columns:
        raw_df["NotificationStatus"] = ""
    if "NotificationChannel" not in raw_df.columns:
        raw_df["NotificationChannel"] = ""
    if "NotificationMessageId" not in raw_df.columns:
        raw_df["NotificationMessageId"] = ""
    if "RoutedToClinicianOn" not in raw_df.columns:
        raw_df["RoutedToClinicianOn"] = pd.NaT
    if "RoutedToClinicianNote" not in raw_df.columns:
        raw_df["RoutedToClinicianNote"] = ""

    mask = raw_df[member_col].astype(str).isin([str(m) for m in member_ids])
    raw_df.loc[mask, "NotificationSentOn"] = ts_date
    raw_df.loc[mask, "NotificationSentAt"] = ts_full

    # Optional per-member metadata (channel, message id, status, etc.)
    if details_by_member_id:
        for mid, details in details_by_member_id.items():
            m = raw_df[member_col].astype(str) == str(mid)
            if not m.any():
                continue
            if "NotificationStatus" in details and details["NotificationStatus"] is not None:
                raw_df.loc[m, "NotificationStatus"] = str(details["NotificationStatus"])
            if "NotificationChannel" in details and details["NotificationChannel"] is not None:
                raw_df.loc[m, "NotificationChannel"] = str(details["NotificationChannel"])
            if "NotificationMessageId" in details and details["NotificationMessageId"] is not None:
                raw_df.loc[m, "NotificationMessageId"] = str(details["NotificationMessageId"])

    raw_df.to_excel(EXCEL_PATH, index=False)

    # refresh cache
    _cached_df = compute_fields(raw_df.copy())
    _last_mtime = os.path.getmtime(EXCEL_PATH)


def update_appreciation_sent_on(
    member_ids: Iterable[str],
    timestamp: Optional[datetime] = None,
) -> None:
    """
    Update AppreciationSentOn column for provided Member IDs (governance: one appreciation per patient per day).
    """
    global _cached_df, _last_mtime

    member_ids = list(member_ids)
    if not member_ids:
        return

    if not os.path.exists(EXCEL_PATH):
        return  # No-op when Excel not configured (e.g. first run)

    ts = pd.Timestamp(timestamp if timestamp is not None else pd.Timestamp.now()).normalize()

    raw_df = pd.read_excel(EXCEL_PATH, engine="openpyxl")
    raw_df.columns = [c.strip() for c in raw_df.columns]

    member_col = "Member ID" if "Member ID" in raw_df.columns else ("Member_ID" if "Member_ID" in raw_df.columns else None)
    if member_col is None:
        raise KeyError("Member ID column not found in the Excel sheet.")

    if "AppreciationSentOn" not in raw_df.columns:
        raw_df["AppreciationSentOn"] = pd.NaT

    mask = raw_df[member_col].astype(str).isin([str(m) for m in member_ids])
    raw_df.loc[mask, "AppreciationSentOn"] = ts

    raw_df.to_excel(EXCEL_PATH, index=False)
    _cached_df = compute_fields(raw_df.copy())
    _last_mtime = os.path.getmtime(EXCEL_PATH)